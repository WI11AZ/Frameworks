"""
Script pour mettre à jour les IDs des abilities dans dcwf_data.json
à partir de DCWF2025.xlsx

Règles:
- Pour chaque page (feuille) commençant par "("
- Les digits 5, 6 et 7 du nom de la page correspondent à l'OPM ID
- Récupérer l'ID de la colonne A si les premiers digits de la colonne C correspondent à "Ability"
"""

import openpyxl
import json
import re
from pathlib import Path

# Chemins des fichiers
excel_path = Path("KSAT25/DCWF2025.xlsx")
json_path = Path("KSAT 2025 FINAL/DCWF 2025/dcwf_data.json")

def extract_opm_id_from_sheet_name(sheet_name):
    """
    Extrait l'OPM ID des digits 5, 6 et 7 du nom de la feuille.
    Exemple: "(IT-441) Net Ops Specialist" -> "441"
    Le format est (XX-XXX) où XXX sont les digits 5, 6, 7
    """
    # Chercher le pattern (XX-XXX) où XXX est l'OPM ID
    match = re.search(r'\([^-]+-(\d{3})\)', sheet_name)
    if match:
        return match.group(1)
    return None

def is_ability_row(cell_c_value):
    """
    Vérifie si les premiers caractères de la colonne C correspondent à "Ability"
    """
    if not cell_c_value:
        return False
    
    cell_str = str(cell_c_value).strip()
    # Vérifier si ça commence par "Ability" (insensible à la casse)
    return cell_str.lower().startswith('ability')

def update_ability_ids():
    """
    Met à jour les IDs des abilities dans le JSON à partir de l'Excel
    """
    print(f"[*] Ouverture du fichier Excel: {excel_path}")
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    
    # Charger le JSON
    print(f"[*] Chargement du JSON: {json_path}")
    with open(json_path, 'r', encoding='utf-8') as f:
        json_data = json.load(f)
    
    # Créer un dictionnaire pour accéder rapidement aux entrées par opm_id et description
    # Structure: {opm_id: {description: index}}
    entries_by_opm = {}
    for idx, entry in enumerate(json_data):
        opm_id = str(entry.get('opm_id', ''))
        if opm_id not in entries_by_opm:
            entries_by_opm[opm_id] = {}
        # Utiliser la description comme clé (normalisée)
        desc = str(entry.get('description', '')).strip().lower()
        if desc:
            entries_by_opm[opm_id][desc] = idx
    
    print(f"[*] {len(json_data)} entrées chargées dans le JSON")
    print(f"[*] {len(entries_by_opm)} OPM IDs uniques trouvés")
    
    # Compteurs
    total_abilities_found = 0
    total_abilities_updated = 0
    abilities_not_found = []
    
    # Parcourir toutes les feuilles
    print("\n[*] Parcours des feuilles Excel...")
    for sheet_name in wb.sheetnames:
        # Vérifier si la feuille commence par "("
        if not sheet_name.startswith('('):
            continue
        
        # Extraire l'OPM ID
        opm_id = extract_opm_id_from_sheet_name(sheet_name)
        if not opm_id:
            print(f"  [!] Impossible d'extraire l'OPM ID de la feuille: {sheet_name}")
            continue
        
        print(f"\n[*] Traitement de la feuille: {sheet_name} (OPM ID: {opm_id})")
        
        ws = wb[sheet_name]
        
        # Parcourir les lignes de la feuille
        abilities_in_sheet = 0
        for row_idx, row in enumerate(ws.iter_rows(min_row=1, values_only=False), start=1):
            # Colonne A (index 0)
            cell_a = row[0].value if len(row) > 0 else None
            # Colonne C (index 2)
            cell_c = row[2].value if len(row) > 2 else None
            
            # Vérifier si c'est une ligne Ability
            if not is_ability_row(cell_c):
                continue
            
            abilities_in_sheet += 1
            total_abilities_found += 1
            
            # Récupérer l'ID de la colonne A
            ability_id = None
            if cell_a:
                ability_id = str(cell_a).strip()
            
            if not ability_id:
                print(f"  [!] Ligne {row_idx}: Pas d'ID trouvé dans la colonne A")
                continue
            
            # Récupérer la description de la colonne C
            description = str(cell_c).strip()
            # Garder la description complète pour la correspondance
            
            # Chercher l'entrée correspondante dans le JSON
            if opm_id not in entries_by_opm:
                abilities_not_found.append({
                    'opm_id': opm_id,
                    'id': ability_id,
                    'description': description,
                    'sheet': sheet_name,
                    'row': row_idx
                })
                print(f"  [!] OPM ID {opm_id} non trouvé dans le JSON")
                continue
            
            # Chercher par description (normalisée)
            desc_normalized = description.lower()
            found = False
            
            # Chercher d'abord une correspondance exacte
            if desc_normalized in entries_by_opm[opm_id]:
                idx = entries_by_opm[opm_id][desc_normalized]
                # Vérifier que c'est bien une ability
                if json_data[idx].get('ksat', '').upper() == 'A':
                    json_data[idx]['id'] = ability_id
                    total_abilities_updated += 1
                    found = True
                    print(f"  [✓] Ligne {row_idx}: ID mis à jour {ability_id} pour '{description[:50]}...'")
            
            # Si pas trouvé, chercher une correspondance partielle
            if not found:
                for desc_key, idx in entries_by_opm[opm_id].items():
                    # Vérifier que c'est une ability
                    if json_data[idx].get('ksat', '').upper() != 'A':
                        continue
                    
                    # Vérifier si la description correspond (partiellement)
                    if desc_normalized in desc_key or desc_key in desc_normalized:
                        json_data[idx]['id'] = ability_id
                        total_abilities_updated += 1
                        found = True
                        print(f"  [✓] Ligne {row_idx}: ID mis à jour {ability_id} (correspondance partielle)")
                        break
            
            if not found:
                abilities_not_found.append({
                    'opm_id': opm_id,
                    'id': ability_id,
                    'description': description,
                    'sheet': sheet_name,
                    'row': row_idx
                })
                print(f"  [!] Ligne {row_idx}: Description non trouvée dans le JSON pour '{description[:50]}...'")
        
        print(f"  [*] {abilities_in_sheet} abilities trouvées dans cette feuille")
    
    # Sauvegarder le JSON mis à jour
    print(f"\n[*] Sauvegarde du JSON mis à jour...")
    with open(json_path, 'w', encoding='utf-8') as f:
        json.dump(json_data, f, indent=2, ensure_ascii=False)
    
    # Résumé
    print(f"\n{'='*60}")
    print(f"RÉSUMÉ:")
    print(f"  - Abilities trouvées dans l'Excel: {total_abilities_found}")
    print(f"  - Abilities mises à jour dans le JSON: {total_abilities_updated}")
    print(f"  - Abilities non trouvées: {len(abilities_not_found)}")
    
    if abilities_not_found:
        print(f"\n[!] Abilities non trouvées:")
        for item in abilities_not_found[:10]:  # Afficher les 10 premiers
            print(f"  - OPM {item['opm_id']}, ID: {item['id']}, Feuille: {item['sheet']}, Ligne: {item['row']}")
        if len(abilities_not_found) > 10:
            print(f"  ... et {len(abilities_not_found) - 10} autres")
    
    print(f"{'='*60}")

if __name__ == "__main__":
    update_ability_ids()

