import json
from openpyxl import load_workbook

# Load Excel file
xlsx_path = "(U)+2025-07-25+DCWF+Work+Role+Tool_v5.1.xlsx"
wb = load_workbook(filename=xlsx_path, data_only=True)

# Initialize JSON skeleton
nice_json = {
    "documents": [
        {
            "doc_identifier": "SP_800_181_rev_1",
            "name": "Workforce Framework for Cybersecurity (NICE Framework) (NIST SP 800-181 Rev 1)",
            "version": "2.0.0",
            "website": "https://csrc.nist.gov/pubs/sp/800/181/r1/final"
        }
    ],
    "relationship_types": [
        {
            "relationship_identifier": "projection",
            "description": "Represents a relationship between two elements."
        }
    ],
    "elements": [],
    "relationships": []
}

# Helper: add an element only if not already present
def add_element(element):
    existing_ids = {e["element_identifier"] for e in nice_json["elements"]}
    if element["element_identifier"] not in existing_ids:
        nice_json["elements"].append(element)

# ------------- STEP 1: Process Tasks and KSAs sheet -------------
task_ksa_sheet = wb.worksheets[2]  # First sheet
for row in task_ksa_sheet.iter_rows(min_row=2, values_only=True):
    if not row or not row[0]:
        continue
    element_type = str(row[3]).strip().lower()   # e.g., "task", "skill", "knowledge"
    element_identifier = str(row[0]).strip()
    title = "" #row[2].strip() if row[2] else ""
    text = row[4].strip() if row[3] else ""
    add_element({
        "element_type": element_type,
        "element_identifier": element_identifier,
        "title": title,
        "text": text,
        "doc_identifier": "SP_800_181_rev_1"
    })

# ------------- STEP 2: Process Work Roles sheet -------------
workrole_sheet = wb.worksheets[3]  # Second sheet
for row in workrole_sheet.iter_rows(min_row=3, values_only=True):
    if not row or not row[4]:
        continue
    dcwf_id = str(row[4]).strip()       # DCWF ID
    ncwf_id = str(row[5]).strip() if row[5] else ""      # NCWF ID
    title = row[3].strip() if row[3] else ""
    text = str(row[6]).strip() if row[6] else ""
    # print(f"dcwf id: {dcwf_id};   ncwf id: {ncwf_id};   title: {title};   text: {text}")
    # Add work role element using DCWF ID
    add_element({
        "element_type": "work_role",
        "element_identifier": dcwf_id,
        "title": title,
        "text": text,
        "doc_identifier": "SP_800_181_rev_1"
    })

    # Create projection relationship from DCWF -> NCWF
    if ncwf_id:
        nice_json["relationships"].append({
            "source_element_identifier": dcwf_id,
            "source_doc_identifier": "SP_800_181_rev_1",
            "dest_element_identifier": ncwf_id,
            "dest_doc_identifier": "SP_800_181_rev_1",
            "relationship_identifier": "projection",
            "provenance_doc_identifier": "SP_800_181_rev_1"
        })

# ------------- STEP 3: Process Individual Work Role Sheets -------------
for sheet in wb.worksheets[5:]:
    sheet_name = sheet.title
    workrole_id = sheet_name.split(")")[0].split("-")[1]  # Assuming sheet name matches DCWF ID
    print (f"sheet {sheet_name } -> workrole {workrole_id} identified")

    for row in sheet.iter_rows(min_row=7, values_only=True):
        if not row or not row[0]:
            continue
        dcwf_id = str(row[0]).strip()
        if not dcwf_id:
            continue
        core_or_additional = str(row[3]).strip() if row[3] else "A"
        # print (f"workrole {workrole_id} - dcwf {dcwf_id} - core/additional {core_or_additional}")
        # Create relationship: workrole_id --> element_id
        nice_json["relationships"].append({
            "source_element_identifier": workrole_id,
            "source_doc_identifier": "SP_800_181_rev_1",
            "dest_element_identifier": dcwf_id,
            "dest_doc_identifier": "SP_800_181_rev_1",
            "relationship_identifier": "projection",
            "provenance_doc_identifier": "SP_800_181_rev_1",
            "core_or_additional": core_or_additional
        })

# ------------- Save Output JSON -------------
with open("(U)+2025-07-25+DCWF+Work+Role+Tool_v5.1.json", "w", encoding="utf-8") as f:
    json.dump(nice_json, f, indent=4)

print("✅ Export complete: (U)+2025-07-25+DCWF+Work+Role+Tool_v5.1.json")
