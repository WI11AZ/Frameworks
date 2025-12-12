"""
Script pour extraire les données de l'Excel et mettre à jour output.json
Règles:
- Ligne 1, colonnes B à CD: OPM_IDs
- Lignes 2-11: Primaires (4 derniers digits en majuscule si cellule non vide)
- Lignes 24-26: Primaires (4 premiers digits en majuscule si cellule non vide)
- Lignes 24-26: Niveaux entre parenthèses pour certains codes
"""

import openpyxl
import json
from pathlib import Path
import re

# Chemins des fichiers
excel_path = Path(r"c:\Users\Godmode\Desktop\Frameworks-master\FORSAPINNOEL\2025-1112__MEGA_SFIA.xlsx")
output_json_path = Path(r"c:\Users\Godmode\Desktop\Frameworks-master\FORSAPINNOEL\output.json")

def extract_skill_code_and_levels(cell_value, is_lines_24_26=False):
    """
    Extrait le code de compétence et les niveaux d'une cellule.
    
    Args:
        cell_value: Valeur de la cellule
        is_lines_24_26: True si c'est les lignes 24-26, False pour lignes 2-11
    
    Returns:
        tuple: (code, levels) où levels est une liste d'entiers ou None
    """
    if not cell_value or cell_value.strip() == "":
        return None, None
    
    cell_value = str(cell_value).strip()
    
    # Extraire le code (4 lettres majuscules)
    if is_lines_24_26:
        # 4 premiers digits en majuscules
        code_match = re.match(r'^([A-Z]{4})', cell_value)
    else:
        # 4 derniers digits en majuscules
        code_match = re.search(r'([A-Z]{4})$', cell_value)
    
    if not code_match:
        return None, None
    
    code = code_match.group(1)
    
    # Extraire les niveaux entre parenthèses (seulement pour lignes 24-26)
    levels = None
    if is_lines_24_26:
        # Chercher un pattern comme (4-6) ou (4-5-6) ou (4, 5, 6)
        level_match = re.search(r'\(([0-9\-,\s]+)\)', cell_value)
        if level_match:
            level_str = level_match.group(1)
            # Parser les niveaux
            if '-' in level_str and ',' not in level_str:
                # Format: 4-6 (range)
                parts = level_str.split('-')
                if len(parts) == 2:
                    try:
                        start = int(parts[0].strip())
                        end = int(parts[1].strip())
                        levels = list(range(start, end + 1))
                    except ValueError:
                        pass
            else:
                # Format: 4, 5, 6 ou 4-5-6
                level_str = level_str.replace('-', ',')
                try:
                    levels = [int(x.strip()) for x in level_str.split(',') if x.strip()]
                except ValueError:
                    pass
    
    return code, levels

def main():
    print(f"[*] Ouverture du fichier Excel: {excel_path}")
    wb = openpyxl.load_workbook(excel_path)
    
    # Prendre la première feuille
    ws = wb.active
    print(f"[*] Feuille active: {ws.title}")
    
    # Extraire les OPM_IDs de la ligne 1, colonnes B à CD
    # B = colonne 2, CD = colonne 82 (2 + 80)
    print("\n[*] Extraction des OPM_IDs (ligne 1, colonnes B a CD)...")
    opm_ids = []
    for col in range(2, 83):  # B=2 à CD=82
        cell_value = ws.cell(row=1, column=col).value
        if cell_value:
            opm_ids.append(str(cell_value).strip())
    
    print(f"[OK] {len(opm_ids)} OPM_IDs trouves: {', '.join(opm_ids[:10])}..." if len(opm_ids) > 10 else f"[OK] {len(opm_ids)} OPM_IDs trouves: {', '.join(opm_ids)}")
    
    # Structure pour stocker les données
    data_by_opm = {}
    
    # Pour chaque colonne (OPM_ID)
    for idx, opm_id in enumerate(opm_ids):
        col = idx + 2  # Décalage de 2 car B = colonne 2
        
        primaires_codes = set()
        primaires_with_levels = {}  # {code: [levels]}
        
        # Extraire des lignes 2-11 (Primaires - 4 derniers digits)
        for row in range(2, 12):  # Lignes 2 à 11
            cell_value = ws.cell(row=row, column=col).value
            code, _ = extract_skill_code_and_levels(cell_value, is_lines_24_26=False)
            if code:
                primaires_codes.add(code)
        
        # Extraire des lignes 24-26 (Primaires - 4 premiers digits + niveaux)
        for row in range(24, 27):  # Lignes 24 à 26
            cell_value = ws.cell(row=row, column=col).value
            code, levels = extract_skill_code_and_levels(cell_value, is_lines_24_26=True)
            if code:
                primaires_codes.add(code)
                if levels:
                    primaires_with_levels[code] = levels
        
        # Stocker les données
        if primaires_codes or primaires_with_levels:
            data_by_opm[opm_id] = {
                'codes': sorted(list(primaires_codes)),
                'levels': primaires_with_levels
            }
            print(f"  OPM {opm_id}: {len(primaires_codes)} codes primaires, {len(primaires_with_levels)} avec niveaux")
    
    # Charger le JSON actuel
    print(f"\n[*] Chargement de {output_json_path}...")
    with open(output_json_path, 'r', encoding='utf-8') as f:
        current_data = json.load(f)
    
    # Mettre à jour le JSON avec les nouvelles données
    print("\n[*] Mise a jour des donnees JSON...")
    updated_count = 0
    
    for item in current_data:
        opm_id = item['OPM_ID']
        
        if opm_id in data_by_opm:
            # Mettre à jour les Primaires
            new_primaires = data_by_opm[opm_id]['codes']
            
            # Créer le nouveau format avec niveaux
            if data_by_opm[opm_id]['levels']:
                # Si on a des niveaux, transformer en format dictionnaire
                item['Primaires'] = []
                for code in new_primaires:
                    if code in data_by_opm[opm_id]['levels']:
                        item['Primaires'].append({
                            'code': code,
                            'levels': data_by_opm[opm_id]['levels'][code]
                        })
                    else:
                        # Code sans niveaux
                        item['Primaires'].append(code)
            else:
                # Aucun niveau, garder le format simple
                item['Primaires'] = new_primaires
            
            updated_count += 1
            print(f"  [OK] OPM {opm_id} mis a jour")
    
    # Sauvegarder le JSON mis à jour
    print(f"\n[*] Sauvegarde dans {output_json_path}...")
    with open(output_json_path, 'w', encoding='utf-8') as f:
        json.dump(current_data, f, indent=2, ensure_ascii=False)
    
    print(f"\n[OK] Termine! {updated_count} OPM_IDs mis a jour.")
    
    # Afficher un exemple
    if updated_count > 0:
        print("\n[*] Exemple de mise a jour:")
        example = [item for item in current_data if item['OPM_ID'] in data_by_opm][:2]
        for item in example:
            print(f"\n  OPM_ID: {item['OPM_ID']}")
            print(f"  Primaires: {json.dumps(item['Primaires'], indent=4, ensure_ascii=False)}")

if __name__ == "__main__":
    main()
